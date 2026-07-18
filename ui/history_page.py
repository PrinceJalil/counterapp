import os
import sqlite3

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QFrame, QTableWidget, QTableWidgetItem, QHeaderView,
    QCheckBox, QAbstractItemView, QMenu, QFileDialog,
    QMessageBox, QCalendarWidget, QDialog, QSpinBox, QComboBox
)
from PyQt6.QtCore import Qt, QDate, QTimer, QEvent, QSize
from PyQt6.QtGui import QColor, QCursor, QIcon

from utils.helpers import get_asset_path, format_number
from utils.theme_manager import theme_manager

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "database", "database.db")


# ─────────────────────────────────────────────────────────────────
#  Fixed Calendar Widget
# ─────────────────────────────────────────────────────────────────
class FixedCalendar(QCalendarWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setGridVisible(False)
        self.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)
        QTimer.singleShot(0, self._patch_year_spinbox)

    def _patch_year_spinbox(self):
        spinbox = self.findChild(QSpinBox, "qt_calendar_yearedit")
        if spinbox:
            spinbox.setKeyboardTracking(True)
            spinbox.editingFinished.connect(self._commit_year)
            spinbox.installEventFilter(self)

    def _commit_year(self):
        spinbox = self.findChild(QSpinBox, "qt_calendar_yearedit")
        if spinbox:
            year = spinbox.value()
            cur  = self.selectedDate()
            new_date = QDate(year, cur.month(), min(cur.day(), QDate(year, cur.month(), 1).daysInMonth()))
            self.setSelectedDate(new_date)
            self.setCurrentPage(new_date.year(), new_date.month())

    def eventFilter(self, obj, event):
        if isinstance(obj, QSpinBox) and event.type() == QEvent.Type.KeyPress:
            if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
                self._commit_year()
                return True
        return super().eventFilter(obj, event)


# ─────────────────────────────────────────────────────────────────
#  Date Range Picker Dialog
# ─────────────────────────────────────────────────────────────────
class DateRangeDialog(QDialog):
    def __init__(self, parent=None, start_date=None, end_date=None):
        super().__init__(parent)
        self.setWindowTitle("Date Filter")
        self.setModal(True)
        self.setFixedSize(660, 340)
        self._date_was_set = False
        
        is_light = theme_manager.is_light_mode
        bg_col = "#ffffff" if is_light else "#1e1f25"
        text_col = "#111318" if is_light else "#e2e2e9"
        cal_bg = "#f4f5f8" if is_light else "#282a2f"
        hover_bg = "#e2e4e9" if is_light else "#37393f"
        border_col = "#d9dce1" if is_light else "#414752"
        
        self.setStyleSheet(f"""
            QDialog {{ background-color: {bg_col}; color: {text_col}; font-family: 'Segoe UI', Arial; }}
            QLabel  {{ color: {text_col}; font-size: 12px; font-weight: bold; background: transparent; }}
            QCalendarWidget {{ background-color: {cal_bg}; color: {text_col}; border-radius: 8px; }}
            QCalendarWidget QAbstractItemView {{
                background-color: {cal_bg}; color: {text_col};
                selection-background-color: #3D8EF0; selection-color: white;
            }}
            QCalendarWidget QWidget#qt_calendar_navigationbar {{ background-color: {bg_col}; }}
            QCalendarWidget QToolButton {{
                color: {text_col}; background: transparent; font-size: 13px; font-weight: bold;
            }}
            QCalendarWidget QToolButton:hover {{ background-color: {hover_bg}; border-radius: 4px; }}
            QCalendarWidget QSpinBox {{
                color: {text_col}; background-color: {cal_bg}; border: none;
                font-size: 13px; font-weight: bold;
            }}
            QPushButton {{
                background-color: {cal_bg}; color: {text_col}; border: 1px solid {border_col};
                border-radius: 6px; padding: 7px 18px; font-size: 13px; font-weight: bold;
            }}
            QPushButton:hover {{ background-color: {hover_bg}; }}
            #btn-apply {{ background-color: #3D8EF0; color: white; border: none; }}
            #btn-apply:hover {{ background-color: {'#2a73cc' if is_light else '#5b9ff2'}; }}
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        cal_row = QHBoxLayout()

        start_col = QVBoxLayout()
        start_col.addWidget(QLabel("Start date"))
        self.cal_start = FixedCalendar()
        if start_date:
            self.cal_start.setSelectedDate(start_date)
        self.cal_start.clicked.connect(lambda _: self._mark_set())
        start_col.addWidget(self.cal_start)

        end_col = QVBoxLayout()
        end_col.addWidget(QLabel("End date"))
        self.cal_end = FixedCalendar()
        if end_date:
            self.cal_end.setSelectedDate(end_date)
        self.cal_end.clicked.connect(lambda _: self._mark_set())
        end_col.addWidget(self.cal_end)

        cal_row.addLayout(start_col)
        cal_row.addSpacing(15)
        cal_row.addLayout(end_col)
        layout.addLayout(cal_row)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_reset = QPushButton("Reset")
        btn_reset.clicked.connect(self._reset)
        btn_apply = QPushButton("Set")
        btn_apply.setObjectName("btn-apply")
        btn_apply.clicked.connect(self.accept)
        btn_row.addWidget(btn_reset)
        btn_row.addSpacing(10)
        btn_row.addWidget(btn_apply)
        layout.addLayout(btn_row)

    def _mark_set(self):
        self._date_was_set = True

    def _reset(self):
        today = QDate.currentDate()
        self.cal_start.setSelectedDate(today.addMonths(-1))
        self.cal_end.setSelectedDate(today)
        self._date_was_set = False

    def get_dates(self):
        return self.cal_start.selectedDate(), self.cal_end.selectedDate(), self._date_was_set


# ─────────────────────────────────────────────────────────────────
#  History Page
# ─────────────────────────────────────────────────────────────────
class HistoryPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._filter_start: QDate | None = None
        self._filter_end:   QDate | None = None
        self._filter_class: str | None   = None
        self._data: list[dict] = []
        self.current_page = 1
        self.rows_per_page = 50

        self._init_ui()
        self._apply_styles()
        self.load_data()

        self._refresh_timer = QTimer(self)
        self._refresh_timer.setInterval(3000)
        self._refresh_timer.timeout.connect(self._auto_refresh)
        self._refresh_timer.start()
        
        theme_manager.theme_changed.connect(self.apply_theme)

    # ── UI Construction ──────────────────────────────────────────────
    def showEvent(self, event):
        self._auto_refresh()
        super().showEvent(event)
        
    def _init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Header
        header = QFrame()
        header.setObjectName("top-header")
        header.setFixedHeight(64)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(25, 0, 25, 0)

        self.title_lbl = QLabel("History Log")
        is_light = theme_manager.is_light_mode
        self.title_lbl.setStyleSheet(f"font-size: 22px; font-weight: 900; color: {'#111318' if is_light else '#e2e2e9'};")

        self.export_top_btn = QPushButton("⬇  Export Tables")
        self.export_top_btn.setObjectName("btn-primary")
        self.export_top_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.export_top_btn.clicked.connect(self._export_all)

        header_layout.addWidget(self.title_lbl)
        header_layout.addStretch()
        header_layout.addWidget(self.export_top_btn)
        main_layout.addWidget(header)

        # Content
        content_wrapper = QWidget()
        content_layout = QVBoxLayout(content_wrapper)
        content_layout.setContentsMargins(30, 20, 30, 20)
        content_layout.setSpacing(25)

        # Stats card
        stats_frame = QFrame()
        stats_frame.setObjectName("stats-card")
        stats_frame.setFixedSize(280, 120)
        stats_layout = QVBoxLayout(stats_frame)
        stats_layout.setContentsMargins(25, 20, 25, 20)
        self.stats_title = QLabel("TOTAL COUNT")
        is_light = theme_manager.is_light_mode
        self.stats_title.setStyleSheet(f"font-size: 11px; font-weight: 800; color: {'#555e6d' if is_light else '#8b919e'}; letter-spacing: 1.5px;")
        self.stats_value = QLabel("0")
        self.stats_value.setStyleSheet(
            f"font-size: 44px; font-weight: bold; color: {'#119c5b' if is_light else '#35e192'}; letter-spacing: -1px;"
        )
        stats_layout.addWidget(self.stats_title)
        stats_layout.addStretch()
        stats_layout.addWidget(self.stats_value)

        stats_wrap = QHBoxLayout()
        stats_wrap.addWidget(stats_frame)
        stats_wrap.addStretch()

        # Table container
        table_container = QFrame()
        table_container.setObjectName("table-container")
        tc_layout = QVBoxLayout(table_container)
        tc_layout.setContentsMargins(0, 0, 0, 0)
        tc_layout.setSpacing(0)

        # Toolbar
        tb_header = QWidget()
        tb_header.setFixedHeight(65)
        tb_header.setObjectName("tb-header")
        tb_h_layout = QHBoxLayout(tb_header)
        tb_h_layout.setContentsMargins(30, 0, 30, 0)

        # Date Filter Dropdown
        self.date_combo = QComboBox()
        self.date_combo.addItems(["All Dates", "Last 7 Days", "Last 30 Days", "Custom Range"])
        self.date_combo.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.date_combo.setMinimumWidth(120)
        self.date_combo.activated.connect(self._on_date_combo_activated)

        # Class Filter Dropdown
        self.class_combo = QComboBox()
        self.class_combo.addItem("All Classes")
        self.class_combo.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.class_combo.setMinimumWidth(130)
        self.class_combo.activated.connect(self._on_class_combo_activated)
        
        tb_h_layout.addWidget(self.date_combo)
        tb_h_layout.addSpacing(15)
        tb_h_layout.addWidget(self.class_combo)
        
        tb_h_layout.addStretch()

        self.btn_del_sel = QPushButton("  Delete Selected")
        self.btn_del_sel.setIcon(QIcon(get_asset_path("trash.svg")))
        self.btn_del_sel.setObjectName("btn-danger-outline")
        self.btn_del_sel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_del_sel.clicked.connect(self._delete_selected)

        self.btn_exp_sel = QPushButton("⬇  Export Selected")
        self.btn_exp_sel.setObjectName("btn-primary")
        self.btn_exp_sel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_exp_sel.clicked.connect(self._export_selected)

        tb_h_layout.addWidget(self.btn_del_sel)
        tb_h_layout.addSpacing(10)
        tb_h_layout.addWidget(self.btn_exp_sel)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        headers = ["", "  DATE", "CLASS", "TOTAL COUNT", "ACTIONS"]
        for i, text in enumerate(headers):
            item = QTableWidgetItem(text)
            alignment = (
                Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
                if i == 1 else Qt.AlignmentFlag.AlignCenter
            )
            item.setTextAlignment(alignment)
            self.table.setHorizontalHeaderItem(i, item)

        self.header_chk = QCheckBox(self.table.horizontalHeader())
        self.header_chk.setStyleSheet("background: transparent;")
        self.header_chk.setGeometry(17, 12, 16, 16)
        self.header_chk.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.header_chk.stateChanged.connect(self._toggle_all_checkboxes)

        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 54)
        self.table.setColumnWidth(4, 90)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setShowGrid(False)
        self.table.verticalHeader().setVisible(False)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.setAlternatingRowColors(True)

        self._checkboxes: list[QCheckBox] = []

        tc_layout.addWidget(tb_header)
        tc_layout.addWidget(self.table)
        
        # Pagination UI
        self.btn_prev = QPushButton("◀ Prev")
        self.btn_prev.setObjectName("btn-outline")
        self.btn_prev.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_prev.clicked.connect(self._prev_page)

        self.lbl_page = QLabel("Page 1 of 1")
        is_light = theme_manager.is_light_mode
        self.lbl_page.setStyleSheet(f"color: {'#555e6d' if is_light else '#8b919e'}; font-size: 13px; font-weight: bold;")
        self.lbl_page.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.btn_next = QPushButton("Next ▶")
        self.btn_next.setObjectName("btn-outline")
        self.btn_next.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_next.clicked.connect(self._next_page)

        pag_layout = QHBoxLayout()
        pag_layout.addStretch()
        pag_layout.addWidget(self.btn_prev)
        pag_layout.addSpacing(15)
        pag_layout.addWidget(self.lbl_page)
        pag_layout.addSpacing(15)
        pag_layout.addWidget(self.btn_next)
        pag_layout.addStretch()

        pag_container = QWidget()
        pag_container.setLayout(pag_layout)
        pag_container.setFixedHeight(50)
        tc_layout.addWidget(pag_container)

        content_layout.addLayout(stats_wrap)
        content_layout.addWidget(table_container)
        main_layout.addWidget(content_wrapper, stretch=1)
        
        self._populate_class_combo()

    def _populate_class_combo(self):
        if not os.path.exists(DB_PATH):
            return
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT class_name FROM history_logs ORDER BY class_name")
        rows = cursor.fetchall()
        conn.close()
        
        current_text = self.class_combo.currentText()
        self.class_combo.blockSignals(True)
        self.class_combo.clear()
        self.class_combo.addItem("All Classes")
        for row in rows:
            self.class_combo.addItem(row[0])
            
        index = self.class_combo.findText(current_text)
        if index >= 0:
            self.class_combo.setCurrentIndex(index)
        self.class_combo.blockSignals(False)

    # ── Database — Read ──────────────────────────────────────────────
    def _fetch_from_db(self) -> list[dict]:
        if not os.path.exists(DB_PATH):
            return []
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        query = """
            SELECT DATE(timestamp) AS date, class_name, SUM(count_in) AS total_count
            FROM history_logs WHERE 1=1
        """
        params = []
        
        if hasattr(self, 'date_combo'):
            idx = self.date_combo.currentIndex()
            if idx == 1: # Last 7 Days
                query += " AND DATE(timestamp) >= DATE('now', '-6 days')"
            elif idx == 2: # Last 30 Days
                query += " AND DATE(timestamp) >= DATE('now', '-29 days')"
            elif idx == 3 and self._filter_start and self._filter_end: # Custom Range
                query += " AND DATE(timestamp) BETWEEN ? AND ?"
                params += [
                    self._filter_start.toString("yyyy-MM-dd"),
                    self._filter_end.toString("yyyy-MM-dd"),
                ]

        if self._filter_class:
            query += " AND class_name = ?"
            params.append(self._filter_class)
            
        query += " GROUP BY DATE(timestamp), class_name ORDER BY date DESC, class_name"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        return [{"date": r[0], "class_name": r[1], "total_count": r[2]} for r in rows]

    # ── Database — Delete ────────────────────────────────────────────
    def _delete_rows_from_db(self, rows: list[dict]):
        if not rows or not os.path.exists(DB_PATH):
            return
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        for row in rows:
            cursor.execute(
                "DELETE FROM history_logs WHERE DATE(timestamp) = ? AND class_name = ?",
                (row["date"], row["class_name"])
            )
        conn.commit()
        conn.close()

    # ── Load / Refresh ───────────────────────────────────────────────
    def load_data(self):
        self._populate_class_combo()
        self._data = self._fetch_from_db()
        self.current_page = 1
        self._rebuild_table()

    def _auto_refresh(self):
        if not self.isVisible():
            return
        self._populate_class_combo()
        new_data = self._fetch_from_db()
        if new_data == self._data:
            return
            
        start_idx = (self.current_page - 1) * self.rows_per_page
        page_data = self._data[start_idx : start_idx + self.rows_per_page]
        
        checked_keys = {
            (page_data[i]["date"], page_data[i]["class_name"])
            for i, chk in enumerate(self._checkboxes)
            if chk.isChecked() and i < len(page_data)
        }
        self._data = new_data
        self._rebuild_table(checked_keys=checked_keys)

    def _rebuild_table(self, checked_keys: set | None = None):
        self.table.setRowCount(0)
        self._checkboxes.clear()
        self.header_chk.setChecked(False)

        total_all = sum(r["total_count"] for r in self._data)
        self.stats_value.setText(format_number(total_all))

        total_pages = max(1, (len(self._data) + self.rows_per_page - 1) // self.rows_per_page)
        if self.current_page > total_pages:
            self.current_page = total_pages
            
        self.lbl_page.setText(f"Page {self.current_page} of {total_pages}")
        self.btn_prev.setEnabled(self.current_page > 1)
        self.btn_next.setEnabled(self.current_page < total_pages)

        start_idx = (self.current_page - 1) * self.rows_per_page
        end_idx = start_idx + self.rows_per_page
        page_data = self._data[start_idx:end_idx]

        for row_idx, row_data in enumerate(page_data):
            self.table.insertRow(row_idx)
            self.table.setRowHeight(row_idx, 52)

            # Col 0 — Checkbox
            chk_widget = QWidget()
            chk_layout = QHBoxLayout(chk_widget)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chk = QCheckBox()
            chk.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            if checked_keys and (row_data["date"], row_data["class_name"]) in checked_keys:
                chk.setChecked(True)
            chk.stateChanged.connect(lambda state, ri=row_idx: self._on_row_check(state, ri))
            chk_layout.addWidget(chk)
            self._checkboxes.append(chk)
            self.table.setCellWidget(row_idx, 0, chk_widget)
            if chk.isChecked():
                self._highlight_row(row_idx, True)

            # Col 1 — Date
            try:
                from datetime import datetime as _dt
                date_str = _dt.strptime(row_data["date"], "%Y-%m-%d").strftime("%d %B %Y")
            except Exception:
                date_str = row_data["date"]
            item_date = QTableWidgetItem(f"  {date_str}")
            item_date.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            item_date.setFlags(item_date.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 1, item_date)

            # Col 2 — Class badge
            cls_widget = QWidget()
            cls_layout = QHBoxLayout(cls_widget)
            cls_layout.setContentsMargins(0, 0, 0, 0)
            cls_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            cls_lbl = QLabel()
            cls_lbl.setObjectName("class-badge")
            cls_lbl.setText(f'<span style="color:#4292f4;">●</span>  {row_data["class_name"]}')
            cls_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            cls_layout.addWidget(cls_lbl)
            self.table.setCellWidget(row_idx, 2, cls_widget)

            # Col 3 — Count
            count_item = QTableWidgetItem(format_number(row_data["total_count"]))
            count_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            count_item.setForeground(QColor("#119c5b" if theme_manager.is_light_mode else "#35e192"))
            count_item.setFlags(count_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            f = count_item.font()
            f.setBold(True)
            f.setPointSize(12)
            count_item.setFont(f)
            self.table.setItem(row_idx, 3, count_item)

            # Col 4 — Delete button
            act_widget = QWidget()
            act_layout = QHBoxLayout(act_widget)
            act_layout.setContentsMargins(0, 0, 0, 0)
            act_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            btn_del = QPushButton()
            btn_del.setIcon(QIcon(get_asset_path("trash.svg")))
            btn_del.setIconSize(QSize(20, 20))
            btn_del.setObjectName("btn-action-delete")
            btn_del.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn_del.setFixedSize(32, 32)
            btn_del.clicked.connect(lambda _, rd=row_data: self._delete_single_row(rd))
            act_layout.addWidget(btn_del)
            self.table.setCellWidget(row_idx, 4, act_widget)

    def _prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self._rebuild_table()

    def _next_page(self):
        total_pages = max(1, (len(self._data) + self.rows_per_page - 1) // self.rows_per_page)
        if self.current_page < total_pages:
            self.current_page += 1
            self._rebuild_table()

    # ── Row highlight ────────────────────────────────────────────────
    def _on_row_check(self, state, row_idx: int):
        self._highlight_row(row_idx, state != 0)

    def _highlight_row(self, row_idx: int, highlight: bool):
        is_light = theme_manager.is_light_mode
        bg_col = QColor("#d2e3fc") if is_light else QColor("#1d3150")
        bg = bg_col if highlight else QColor(0, 0, 0, 0)
        for col in [1, 3]:
            item = self.table.item(row_idx, col)
            if item:
                if highlight:
                    item.setBackground(bg)
                else:
                    item.setData(Qt.ItemDataRole.BackgroundRole, None)

    # ── Actions ──────────────────────────────────────────────────────
    def _delete_single_row(self, row_data: dict):
        reply = QMessageBox.question(
            self, "Konfirmasi Hapus",
            f"Hapus semua data untuk <b>{row_data['class_name']}</b> "
            f"pada <b>{row_data['date']}</b>?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._delete_rows_from_db([row_data])
            self.load_data()

    def _get_checked_rows(self) -> list[dict]:
        start_idx = (self.current_page - 1) * self.rows_per_page
        page_data = self._data[start_idx : start_idx + self.rows_per_page]
        return [
            page_data[i]
            for i, chk in enumerate(self._checkboxes)
            if chk.isChecked() and i < len(page_data)
        ]

    def _delete_selected(self):
        rows = self._get_checked_rows()
        if not rows:
            QMessageBox.information(self, "Tidak Ada Pilihan", "Pilih baris terlebih dahulu.")
            return
        reply = QMessageBox.question(
            self, "Konfirmasi Hapus",
            f"Hapus {len(rows)} baris data yang dipilih?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._delete_rows_from_db(rows)
            self.load_data()


    # Export
    def _export_rows_to_excel(self, rows: list[dict]):
        if not rows:
            QMessageBox.information(self, "Tidak Ada Data", "Tidak ada data untuk diekspor.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Simpan Excel", "history_export.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "History Log"

            header_fill   = PatternFill("solid", fgColor="1E3A5F")
            header_font   = Font(bold=True, color="FFFFFF", size=11, name="Segoe UI")
            header_align  = Alignment(horizontal="center", vertical="center")
            header_border = Border(bottom=Side(style="thin", color="3D8EF0"))
            col_headers   = ["No", "Date", "Class", "Total Count"]
            col_widths    = [6, 20, 22, 16]

            for col_idx, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = header_align
                cell.border    = header_border
                ws.column_dimensions[get_column_letter(col_idx)].width = w
            ws.row_dimensions[1].height = 24

            count_font   = Font(name="Segoe UI", bold=True, size=10, color="000000")
            base_font    = Font(name="Segoe UI", size=10, color="000000")
            center_align = Alignment(horizontal="center", vertical="center")
            left_align   = Alignment(horizontal="left",   vertical="center")
            white_fill   = PatternFill("solid", fgColor="FFFFFF")

            for i, row in enumerate(rows):
                r = i + 2
                try:
                    from datetime import datetime as _dt
                    date_str = _dt.strptime(row["date"], "%Y-%m-%d").strftime("%d %B %Y")
                except Exception:
                    date_str = row["date"]
                
                c1 = ws.cell(r, 1, i + 1)
                c1.alignment = center_align
                c1.font = base_font
                c1.fill = white_fill
                
                c2 = ws.cell(r, 2, date_str)
                c2.alignment = left_align
                c2.font = base_font
                c2.fill = white_fill
                
                c3 = ws.cell(r, 3, row["class_name"])
                c3.alignment = center_align
                c3.font = base_font
                c3.fill = white_fill
                
                count_cell = ws.cell(r, 4, row["total_count"])
                count_cell.font      = count_font
                count_cell.alignment = center_align
                count_cell.fill      = white_fill
                
                ws.row_dimensions[r].height = 20

            ws.freeze_panes = "A2"
            wb.save(path)
            QMessageBox.information(self, "Berhasil", f"Data berhasil diekspor ke:\n{path}")
        except ImportError:
            QMessageBox.critical(
                self, "Library Tidak Ditemukan",
                "Modul 'openpyxl' belum terinstall.\n\nJalankan: pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal mengekspor: {e}")

    def _export_all(self):
        self._export_rows_to_excel(self._data)

    def _export_selected(self):
        rows = self._get_checked_rows()
        if not rows:
            QMessageBox.information(self, "Tidak Ada Pilihan", "Pilih baris terlebih dahulu.")
            return
        self._export_rows_to_excel(rows)


    # Filters
    def _on_date_combo_activated(self, index):
        if index == 3: # Custom Range
            dlg = DateRangeDialog(self, self._filter_start, self._filter_end)
            if dlg.exec():
                s, e, was_set = dlg.get_dates()
                if was_set and s <= e:
                    self._filter_start = s
                    self._filter_end   = e
                    custom_text = f"Custom Range"
                    
                    self.date_combo.blockSignals(True)
                    if self.date_combo.count() == 4:
                        self.date_combo.setItemText(3, custom_text)
                    self.date_combo.setCurrentIndex(3)
                    self.date_combo.blockSignals(False)
                    self.load_data()
                elif not was_set:
                    self._filter_start = None
                    self._filter_end = None
                    self.date_combo.setItemText(3, "Custom Range")
                    self.date_combo.setCurrentIndex(0) # Revert to All Dates
                    self.load_data()
                else:
                    QMessageBox.warning(self, "Tanggal Salah", "Tanggal mulai harus sebelum tanggal akhir.")
                    self.date_combo.setCurrentIndex(0)
                    self.date_combo.setItemText(3, "Custom Range")
                    self.load_data()
            else:
                self._sync_date_dropdown()
        else:
            self._filter_start = None
            self._filter_end = None
            self.date_combo.setItemText(3, "Custom Range")
            self.load_data()

    def _sync_date_dropdown(self):
        if self._filter_start and self._filter_end:
            self.date_combo.setCurrentIndex(3)
        else:
            pass

    def _on_class_combo_activated(self, index):
        text = self.class_combo.itemText(index)
        if text == "All Classes":
            self._filter_class = None
        else:
            self._filter_class = text
        self.load_data()

    # Checkbox helpers
    def _toggle_all_checkboxes(self, state):
        is_checked = (state != 0)
        for chk in self._checkboxes:
            if chk.isChecked() != is_checked:
                chk.setChecked(is_checked)

    # Styles
    def apply_theme(self, is_light: bool):
        self.title_lbl.setStyleSheet(f"font-size: 22px; font-weight: 900; color: {'#111318' if is_light else '#e2e2e9'};")
        self.stats_title.setStyleSheet(f"font-size: 11px; font-weight: 800; color: {'#555e6d' if is_light else '#8b919e'}; letter-spacing: 1.5px;")
        self.stats_value.setStyleSheet(f"font-size: 44px; font-weight: bold; color: {'#119c5b' if is_light else '#35e192'}; letter-spacing: -1px;")
        self.lbl_page.setStyleSheet(f"color: {'#555e6d' if is_light else '#8b919e'}; font-size: 13px; font-weight: bold;")
        self._apply_styles()
        self._rebuild_table() # Redraw table items to update highlighting and foreground colors

    def _apply_styles(self):
        is_light = theme_manager.is_light_mode
        
        bg_main = "#f4f5f8" if is_light else "#111318"
        border_col = "#d9dce1" if is_light else "#282a2f"
        text_col = "#111318" if is_light else "#e2e2e9"
        input_bg = "#ffffff" if is_light else "#1a1b21"
        input_border = "#c8cbd2" if is_light else "#33353a"
        dropdown_bg = "#e2e4e9" if is_light else "#282a2f"
        list_bg = "#ffffff" if is_light else "#1e1f25"
        card_bg = "#ffffff" if is_light else "#0c0e13"
        table_bg = "#ffffff" if is_light else "#1e1f25"
        danger_bg = "rgba(217, 48, 37, 0.1)" if is_light else "rgba(147, 0, 10, 0.2)"
        danger_hover = "rgba(217, 48, 37, 0.8)" if is_light else "rgba(200, 30, 30, 0.8)"
        danger_text = "#d93025" if is_light else "white"
        row_sel = "#d2e3fc" if is_light else "#1d3150"
        
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet(f"""
            HistoryPage {{ background-color: {bg_main}; }}
            #top-header {{ background-color: {bg_main}; border-bottom: 1px solid {border_col}; }}
            QWidget {{ color: {text_col}; font-family: 'Segoe UI', 'Inter', Arial, sans-serif; }}
            #btn-primary {{
                background-color: #3D8EF0; color: white; border-radius: 6px;
                padding: 8px 18px; font-weight: bold; font-size: 13px; border: none;
            }}
            #btn-primary:hover {{ background-color: {'#2a73cc' if is_light else '#5b9ff2'}; }}
            
            QComboBox {{
                background-color: {input_bg};
                color: {'#555e6d' if is_light else '#8b919e'};
                border: 1px solid {input_border};
                border-radius: 8px;
                padding: 6px 30px 6px 14px;
                font-size: 12px;
                font-weight: bold;
            }}
            QComboBox:hover {{
                background-color: {dropdown_bg};
                color: {text_col};
            }}
            QComboBox::drop-down {{
                border: none;
                width: 30px;
            }}
            QComboBox::down-arrow {{
                image: url('data:image/svg+xml;utf8,<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="%23{'555e6d' if is_light else '8b919e'}" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><polyline points="6 9 12 15 18 9"></polyline></svg>');
            }}
            QComboBox:hover::down-arrow {{
                image: url('data:image/svg+xml;utf8,<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="%23{'111318' if is_light else 'e2e2e9'}" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><polyline points="6 9 12 15 18 9"></polyline></svg>');
            }}
            QComboBox QAbstractItemView {{
                background-color: {list_bg};
                color: {text_col};
                selection-background-color: #3D8EF0;
                border: 1px solid {input_border};
                outline: none;
            }}

            #stats-card {{ background-color: {card_bg}; border-radius: 16px; border: 1px solid {border_col}; }}
            #table-container {{
                background-color: {table_bg}; border-radius: 12px; border: 1px solid {border_col};
            }}
            #tb-header {{ background-color: transparent; border-bottom: 1px solid {input_border}; }}
            #btn-danger-outline {{
                background-color: {danger_hover}; color: white;
                border-radius: 6px; padding: 8px 18px; font-size: 13px; font-weight: bold; border: none;
            }}
            #btn-danger-outline:hover {{ background-color: #a82626; color: white; }}
            #btn-outline {{
                background-color: {dropdown_bg}; color: {text_col}; border: 1px solid {input_border};
                border-radius: 6px; padding: 6px 14px; font-size: 12px; font-weight: 600;
            }}
            #btn-outline:hover {{ background-color: {input_border}; }}
            QTableWidget {{
                background-color: transparent; border: none;
                gridline-color: transparent; color: {text_col}; font-size: 13px; outline: none;
            }}
            QTableWidget::item {{ border-bottom: 1px solid {'rgba(0,0,0,0.05)' if is_light else 'rgba(255,255,255,0.03)'}; padding: 0px 4px; }}
            QTableWidget::item:selected {{ background-color: {row_sel}; }}
            QHeaderView {{
                background-color: {table_bg}; border: none;
            }}
            QHeaderView::section {{
                background-color: {dropdown_bg}; color: {'#555e6d' if is_light else '#8b919e'};
                font-size: 11px; font-weight: bold; letter-spacing: 1px;
                padding: 12px 10px; border: none;
            }}
            #class-badge {{
                background-color: {dropdown_bg}; color: {text_col}; border-radius: 6px;
                padding: 5px 10px; font-size: 12px; font-weight: bold;
            }}
            #btn-action-delete {{
                background-color: {danger_hover}; color: white; border-radius: 4px; padding: 4px 8px; border: none;
            }}
            #btn-action-delete:hover {{ background-color: #a82626; color: white; }}
            QCheckBox::indicator {{
                width: 17px; height: 17px; border: 1.5px solid {border_col};
                border-radius: 4px; background-color: {input_bg};
            }}
            QCheckBox::indicator:checked {{ background-color: #3D8EF0; border: 1.5px solid #3D8EF0; }}
            QCheckBox::indicator:hover {{ border-color: #3D8EF0; }}
            QTableWidget::item:alternate {{ background-color: {'#f4f5f8' if is_light else '#17181e'}; }}
            QScrollBar:vertical {{ border: none; background: {bg_main}; width: 8px; margin: 0px; }}
            QScrollBar::handle:vertical {{ background: {input_border}; min-height: 20px; border-radius: 4px; }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ border: none; background: none; height: 0px; }}
        """)
